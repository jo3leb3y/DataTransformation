from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json

work_book = load_workbook(filename=r"") #excel workbook path
work_sheet = work_book.active

my_list = []
for work_sheet in work_book:

    last_row = len(list(work_sheet.rows))
    last_column = len(list(work_sheet.columns))

    for row in range(1, last_row + 1):
        my_dict = {}
        for column in range(1, last_column + 1):
            column_letter = get_column_letter(column)
            if row > 1:
                my_dict[work_sheet[column_letter + str(1)].value] = work_sheet[column_letter + str(row)].value
        my_list.append(my_dict)
        
data = json.dumps(my_list, sort_keys=True, indent=4)
with open(r'', 'w', encoding='utf-8') as f:         #artworks.txt path
    f.write(data)